from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .models import Case, Step

REQUIRED_SHEETS = {"cases", "steps"}


def load_cases(excel_path: str | Path) -> list[Case]:
    workbook = load_workbook(filename=excel_path, data_only=True)
    missing = REQUIRED_SHEETS.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Excel 缺少 sheet: {', '.join(sorted(missing))}")

    case_sheet = workbook["cases"]
    step_sheet = workbook["steps"]

    case_headers = _headers(case_sheet)
    step_headers = _headers(step_sheet)

    cases_by_id: dict[str, Case] = {}
    for row in case_sheet.iter_rows(min_row=2, values_only=True):
        payload = _row_to_dict(case_headers, row)
        case_id = _to_text(payload.get("case_id"))
        if not case_id:
            continue
        case = Case(
            case_id=case_id,
            title=_to_text(payload.get("title")) or case_id,
            enabled=_to_bool(payload.get("enabled"), default=True),
            start_url=_to_text(payload.get("start_url")),
            tags=_to_text(payload.get("tags")),
            notes=_to_text(payload.get("notes")),
        )
        cases_by_id[case_id] = case

    for row in step_sheet.iter_rows(min_row=2, values_only=True):
        payload = _row_to_dict(step_headers, row)
        case_id = _to_text(payload.get("case_id"))
        if not case_id:
            continue
        case = cases_by_id.get(case_id)
        if case is None:
            raise ValueError(f"steps 中引用了不存在的 case_id: {case_id}")
        step = Step(
            case_id=case_id,
            step_no=_to_int(payload.get("step_no"), default=len(case.steps) + 1),
            action=_to_text(payload.get("action")).strip(),
            locator=_to_text(payload.get("locator")),
            value=_to_text(payload.get("value")),
            timeout_ms=_to_int(payload.get("timeout_ms"), default=5000),
            description=_to_text(payload.get("description")),
        )
        if not step.action:
            continue
        case.steps.append(step)

    cases = sorted(cases_by_id.values(), key=lambda item: item.case_id)
    for case in cases:
        case.steps.sort(key=lambda item: item.step_no)
    return cases


def _headers(sheet) -> list[str]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(cell).strip() if cell is not None else "" for cell in header_row]


def _row_to_dict(headers: list[str], row: tuple[object, ...]) -> dict[str, object]:
    return {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}


def _to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: object, default: int) -> int:
    if value in (None, ""):
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(str(value).strip())


def _to_bool(value: object, default: bool) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "enabled", "on"}:
        return True
    if text in {"0", "false", "no", "n", "disabled", "off"}:
        return False
    return default
