from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

TARGET_URL = "https://dgx.xlook.ai/dgx/93e7ea32-0a24-4ad9-ad46-d3362e027ab7"
OUTPUT_PATH = Path("data/dgx_business_flow_cases.xlsx")


CASES = [
    ["DGX-BIZ-001", "站点入口与 Site Conf. 首屏加载", 1, TARGET_URL, "business-flow,site-conf", "验证首屏、地图、项目名称、导航按钮"],
    ["DGX-BIZ-002", "进入 Surrounding Conf. 并校验周边清单", 1, TARGET_URL, "business-flow,surrounding", "验证周边 tab 和典型建筑数据已经渲染"],
    ["DGX-BIZ-003", "进入 Site Analysis 并校验分析卡片", 1, TARGET_URL, "business-flow,analysis", "验证 Site Information、Plot Detail、Analysis Tools 区块"],
    ["DGX-BIZ-004", "进入 Design Conf. 并校验设计阶段地图", 1, TARGET_URL, "business-flow,design", "验证设计阶段仍能正常进入地图视图"],
    ["DGX-BIZ-005", "进入 Plan Viewing 并校验计划列表状态", 1, TARGET_URL, "business-flow,plan-viewing", "验证 Plan List、Save 按钮、关键词框、生成状态文案"],
    ["DGX-BIZ-006", "端到端贯通五步主流程", 1, TARGET_URL, "business-flow,e2e", "单条用例串起五个阶段，适合作为主流程回归"],
    ["DGX-BIZ-007", "探索性：地图地标文字是否出现", 0, TARGET_URL, "exploratory,map-text", "POI 文字受缩放和瓦片状态影响，默认关闭"],
    ["DGX-BIZ-008", "危险操作预留：Plan Viewing 的 Save", 0, TARGET_URL, "dangerous,save", "Save 可能写远端数据，默认关闭，需人工确认后再开"],
]


STEPS = [
    ["DGX-BIZ-001", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-001", 2, "wait_for_timeout", "", "2000", 2000, "等待首页渲染"],
    ["DGX-BIZ-001", 3, "assert_visible", "text=Site Conf.", "", 3000, "步骤导航处于 Site Conf. 阶段"],
    ["DGX-BIZ-001", 4, "assert_visible", "text=Faya Al Saadiyat", "", 5000, "项目名称/地块名称可见"],
    ["DGX-BIZ-001", 5, "assert_visible", ".leaflet-container", "", 3000, "地图容器可见"],
    ["DGX-BIZ-001", 6, "assert_visible", "role=button[name='Next']", "", 3000, "Next 按钮可见"],
    ["DGX-BIZ-001", 7, "assert_visible", "role=button[name='Back']", "", 3000, "首屏 Back 按钮可见"],
    ["DGX-BIZ-001", 8, "screenshot", "", "biz-001-site-conf", 3000, "保存首屏业务截图"],

    ["DGX-BIZ-002", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-002", 2, "wait_for_timeout", "", "1500", 1500, "等待首页渲染"],
    ["DGX-BIZ-002", 3, "click", "role=button[name='Next']", "", 5000, "进入 Surrounding Conf."],
    ["DGX-BIZ-002", 4, "wait_for_timeout", "", "2000", 2000, "等待周边配置数据渲染"],
    ["DGX-BIZ-002", 5, "assert_visible", "text=Surrounding Buildings", "", 3000, "周边建筑 tab 可见"],
    ["DGX-BIZ-002", 6, "assert_visible", "text=POI", "", 3000, "POI tab 可见"],
    ["DGX-BIZ-002", 7, "assert_visible", "text=Roads", "", 3000, "Roads tab 可见"],
    ["DGX-BIZ-002", 8, "assert_visible", "text=Zones", "", 3000, "Zones tab 可见"],
    ["DGX-BIZ-002", 9, "assert_visible", "text=Park Hyatt Abu Dhabi Hotel", "", 5000, "典型周边建筑记录存在"],
    ["DGX-BIZ-002", 10, "assert_visible", "text=ADNOC", "", 5000, "表格中可见另一条典型记录"],
    ["DGX-BIZ-002", 11, "screenshot", "", "biz-002-surrounding-conf", 3000, "保存周边配置截图"],

    ["DGX-BIZ-003", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-003", 2, "wait_for_timeout", "", "1500", 1500, "等待首页渲染"],
    ["DGX-BIZ-003", 3, "click", "role=button[name='Next']", "", 5000, "进入 Surrounding Conf."],
    ["DGX-BIZ-003", 4, "wait_for_timeout", "", "1200", 1200, "等待页面切换"],
    ["DGX-BIZ-003", 5, "click", "role=button[name='Next']", "", 5000, "进入 Site Analysis"],
    ["DGX-BIZ-003", 6, "wait_for_timeout", "", "2000", 2000, "等待分析卡片渲染"],
    ["DGX-BIZ-003", 7, "assert_visible", "text=Site Information", "", 3000, "站点信息区域可见"],
    ["DGX-BIZ-003", 8, "assert_visible", "text=Project Detail", "", 3000, "Project Detail 按钮可见"],
    ["DGX-BIZ-003", 9, "assert_visible", "text=Site Area", "", 3000, "Site Area 指标可见"],
    ["DGX-BIZ-003", 10, "assert_visible", "text=Plot 1 - Faya Al Saadiyat", "", 3000, "Plot 1 卡片可见"],
    ["DGX-BIZ-003", 11, "assert_visible", "text=Analysis Tools", "", 3000, "Analysis Tools 区域可见"],
    ["DGX-BIZ-003", 12, "assert_visible", "text=Annual Solar Path", "", 3000, "Solar Path 工具可见"],
    ["DGX-BIZ-003", 13, "assert_count_gte", "role=button[name='View Result']", "4", 3000, "至少有 4 个 View Result 按钮"],
    ["DGX-BIZ-003", 14, "assert_count_gte", "text=Analysis completed", "4", 3000, "至少有 4 个分析完成状态"],
    ["DGX-BIZ-003", 15, "screenshot", "", "biz-003-site-analysis", 3000, "保存分析页截图"],

    ["DGX-BIZ-004", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-004", 2, "wait_for_timeout", "", "1500", 1500, "等待首页渲染"],
    ["DGX-BIZ-004", 3, "click", "role=button[name='Next']", "", 5000, "进入 Surrounding Conf."],
    ["DGX-BIZ-004", 4, "wait_for_timeout", "", "1000", 1000, "等待页面切换"],
    ["DGX-BIZ-004", 5, "click", "role=button[name='Next']", "", 5000, "进入 Site Analysis"],
    ["DGX-BIZ-004", 6, "wait_for_timeout", "", "1000", 1000, "等待页面切换"],
    ["DGX-BIZ-004", 7, "click", "role=button[name='Next']", "", 5000, "进入 Design Conf."],
    ["DGX-BIZ-004", 8, "wait_for_timeout", "", "2000", 2000, "等待设计阶段地图渲染"],
    ["DGX-BIZ-004", 9, "assert_visible", "text=Design Conf.", "", 3000, "设计配置步骤高亮存在"],
    ["DGX-BIZ-004", 10, "assert_visible", ".leaflet-container", "", 3000, "设计阶段地图容器可见"],
    ["DGX-BIZ-004", 11, "assert_count_gte", "text=Faya Al Saadiyat", "3", 3000, "设计阶段至少出现 3 个地块标签"],
    ["DGX-BIZ-004", 12, "screenshot", "", "biz-004-design-conf", 3000, "保存设计阶段截图"],

    ["DGX-BIZ-005", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-005", 2, "wait_for_timeout", "", "1500", 1500, "等待首页渲染"],
    ["DGX-BIZ-005", 3, "click", "role=button[name='Next']", "", 5000, "进入 Surrounding Conf."],
    ["DGX-BIZ-005", 4, "wait_for_timeout", "", "800", 800, "等待页面切换"],
    ["DGX-BIZ-005", 5, "click", "role=button[name='Next']", "", 5000, "进入 Site Analysis"],
    ["DGX-BIZ-005", 6, "wait_for_timeout", "", "800", 800, "等待页面切换"],
    ["DGX-BIZ-005", 7, "click", "role=button[name='Next']", "", 5000, "进入 Design Conf."],
    ["DGX-BIZ-005", 8, "wait_for_timeout", "", "800", 800, "等待页面切换"],
    ["DGX-BIZ-005", 9, "click", "role=button[name='Next']", "", 5000, "进入 Plan Viewing"],
    ["DGX-BIZ-005", 10, "wait_for_timeout", "", "2500", 2500, "等待计划列表初始化"],
    ["DGX-BIZ-005", 11, "assert_visible", "text=Plan List", "", 3000, "Plan List 标题可见"],
    ["DGX-BIZ-005", 12, "assert_visible", "role=button[name='Save']", "", 3000, "Save 按钮可见"],
    ["DGX-BIZ-005", 13, "assert_visible", "input[placeholder='DGX Exercise Keywords']", "", 3000, "关键词输入框可见"],
    ["DGX-BIZ-005", 14, "assert_visible", "text=Plan generation queued", "", 5000, "计划生成已排队文案可见"],
    ["DGX-BIZ-005", 15, "screenshot", "", "biz-005-plan-viewing", 3000, "保存计划列表截图"],

    ["DGX-BIZ-006", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-006", 2, "wait_for_timeout", "", "1800", 1800, "等待首页渲染"],
    ["DGX-BIZ-006", 3, "assert_visible", "text=Site Conf.", "", 3000, "确认第 1 步"],
    ["DGX-BIZ-006", 4, "click", "role=button[name='Next']", "", 5000, "前往第 2 步"],
    ["DGX-BIZ-006", 5, "wait_for_timeout", "", "1500", 1500, "等待第 2 步渲染"],
    ["DGX-BIZ-006", 6, "assert_visible", "text=Surrounding Buildings", "", 3000, "确认第 2 步"],
    ["DGX-BIZ-006", 7, "click", "role=button[name='Next']", "", 5000, "前往第 3 步"],
    ["DGX-BIZ-006", 8, "wait_for_timeout", "", "1500", 1500, "等待第 3 步渲染"],
    ["DGX-BIZ-006", 9, "assert_visible", "text=Analysis Tools", "", 3000, "确认第 3 步"],
    ["DGX-BIZ-006", 10, "click", "role=button[name='Next']", "", 5000, "前往第 4 步"],
    ["DGX-BIZ-006", 11, "wait_for_timeout", "", "1500", 1500, "等待第 4 步渲染"],
    ["DGX-BIZ-006", 12, "assert_visible", ".leaflet-container", "", 3000, "确认第 4 步地图"],
    ["DGX-BIZ-006", 13, "click", "role=button[name='Next']", "", 5000, "前往第 5 步"],
    ["DGX-BIZ-006", 14, "wait_for_timeout", "", "2500", 2500, "等待第 5 步渲染"],
    ["DGX-BIZ-006", 15, "assert_visible", "text=Plan List", "", 3000, "确认第 5 步"],
    ["DGX-BIZ-006", 16, "assert_visible", "text=Plan generation queued", "", 5000, "确认计划生成已排队"],
    ["DGX-BIZ-006", 17, "screenshot", "", "biz-006-e2e-plan-viewing", 3000, "保存端到端终点截图"],

    ["DGX-BIZ-007", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-007", 2, "wait_for_timeout", "", "2500", 2500, "等待 POI 文字渲染"],
    ["DGX-BIZ-007", 3, "assert_visible", "text=Faya Al Saadiyat", "", 3000, "地标/地块文字可见"],

    ["DGX-BIZ-008", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"],
    ["DGX-BIZ-008", 2, "wait_for_timeout", "", "1500", 1500, "等待首页渲染"],
    ["DGX-BIZ-008", 3, "click", "role=button[name='Next']", "", 5000, "进入 Surrounding Conf."],
    ["DGX-BIZ-008", 4, "wait_for_timeout", "", "800", 800, "等待页面切换"],
    ["DGX-BIZ-008", 5, "click", "role=button[name='Next']", "", 5000, "进入 Site Analysis"],
    ["DGX-BIZ-008", 6, "wait_for_timeout", "", "800", 800, "等待页面切换"],
    ["DGX-BIZ-008", 7, "click", "role=button[name='Next']", "", 5000, "进入 Design Conf."],
    ["DGX-BIZ-008", 8, "wait_for_timeout", "", "800", 800, "等待页面切换"],
    ["DGX-BIZ-008", 9, "click", "role=button[name='Next']", "", 5000, "进入 Plan Viewing"],
    ["DGX-BIZ-008", 10, "wait_for_timeout", "", "2500", 2500, "等待计划列表初始化"],
    ["DGX-BIZ-008", 11, "assert_visible", "role=button[name='Save']", "", 3000, "Save 按钮可见，但默认不点击"],
]


COVERAGE = [
    ["模块", "用例ID", "覆盖内容", "是否默认启用"],
    ["Site Conf.", "DGX-BIZ-001", "入口、地图、项目名称、Next/Back 状态", "是"],
    ["Surrounding Conf.", "DGX-BIZ-002", "周边 tab、建筑数据清单", "是"],
    ["Site Analysis", "DGX-BIZ-003", "Site Information、Plot、Analysis Tools", "是"],
    ["Design Conf.", "DGX-BIZ-004", "设计阶段地图与地块标签", "是"],
    ["Plan Viewing", "DGX-BIZ-005", "Plan List、Save、关键词框、排队状态", "是"],
    ["E2E 主流程", "DGX-BIZ-006", "1→5 步完整串联", "是"],
    ["探索性", "DGX-BIZ-007", "易波动地图文字", "否"],
    ["危险操作", "DGX-BIZ-008", "Save 前的可见性预检", "否"],
]


NOTES = [
    ["说明"],
    ["这份 Excel 是按当前线上页面实际可见流程拆的第一版业务用例。"],
    ["默认启用的 6 条用例都避免了明显写操作，适合日常回归。"],
    ["DGX-BIZ-007 和 DGX-BIZ-008 默认关闭：前者易波动，后者可能触发远端保存。"],
    ["如果后续确认允许触发 Save / Run Analysis，可以再把对应用例从 0 改成 1。"],
]


def main() -> None:
    workbook = Workbook()

    cases_sheet = workbook.active
    cases_sheet.title = "cases"
    cases_sheet.append(["case_id", "title", "enabled", "start_url", "tags", "notes"])
    for row in CASES:
        cases_sheet.append(row)

    steps_sheet = workbook.create_sheet("steps")
    steps_sheet.append(["case_id", "step_no", "action", "locator", "value", "timeout_ms", "description"])
    for row in STEPS:
        steps_sheet.append(row)

    coverage_sheet = workbook.create_sheet("coverage")
    for row in COVERAGE:
        coverage_sheet.append(row)

    notes_sheet = workbook.create_sheet("notes")
    for row in NOTES:
        notes_sheet.append(row)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"业务流程 Excel 已生成: {OUTPUT_PATH.resolve()}")


if __name__ == "__main__":
    main()
