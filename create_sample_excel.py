from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

TARGET_URL = "https://dgx.xlook.ai/dgx/93e7ea32-0a24-4ad9-ad46-d3362e027ab7"
OUTPUT_PATH = Path("data/cases.xlsx")


def main() -> None:
    workbook = Workbook()

    cases_sheet = workbook.active
    cases_sheet.title = "cases"
    cases_sheet.append(["case_id", "title", "enabled", "start_url", "tags", "notes"])
    cases_sheet.append(["DGX-001", "首页基础可访问", 1, TARGET_URL, "smoke,title", "确认页面可打开且标题/URL正确"])
    cases_sheet.append(["DGX-002", "顶部步骤导航与按钮可见", 1, TARGET_URL, "smoke,header", "确认关键导航文本和 Back/Next 按钮存在"])
    cases_sheet.append(["DGX-003", "地图容器与版权信息可见", 1, TARGET_URL, "smoke,map", "用 Leaflet 容器和版权信息做更稳定的地图 smoke"])
    cases_sheet.append(["DGX-004", "页面 DOM 数量基础检查", 1, TARGET_URL, "smoke,render", "验证页面不是空白壳子，也顺手覆盖计数断言"])
    cases_sheet.append(["DGX-005", "前端错误阈值检查", 0, TARGET_URL, "debug,console", "默认关闭；若第三方资源常报错，可把阈值调大"])
    cases_sheet.append(["DGX-006", "地图 POI 文字探测", 0, TARGET_URL, "exploratory,map-text", "地图上的 POI 文字会随缩放/瓦片状态波动，默认关闭"])

    steps_sheet = workbook.create_sheet("steps")
    steps_sheet.append(["case_id", "step_no", "action", "locator", "value", "timeout_ms", "description"])

    steps_sheet.append(["DGX-001", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"])
    steps_sheet.append(["DGX-001", 2, "wait_for_timeout", "", "1500", 1500, "给首屏一点渲染时间"])
    steps_sheet.append(["DGX-001", 3, "assert_title_contains", "", "DGX Design", 3000, "校验网页标题"])
    steps_sheet.append(["DGX-001", 4, "assert_url_contains", "", "/dgx/", 3000, "校验当前 URL"])
    steps_sheet.append(["DGX-001", 5, "screenshot", "", "home-title-ok", 3000, "保存首屏截图"])

    steps_sheet.append(["DGX-002", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"])
    steps_sheet.append(["DGX-002", 2, "wait_for_selector", "role=button[name='Next']", "visible", 8000, "等待 Next 按钮出现"])
    steps_sheet.append(["DGX-002", 3, "assert_visible", "text=Site Conf.", "", 3000, "当前步骤文字可见"])
    steps_sheet.append(["DGX-002", 4, "assert_visible", "text=Plan Viewing", "", 3000, "末尾步骤文字可见"])
    steps_sheet.append(["DGX-002", 5, "assert_visible", "role=button[name='Back']", "", 3000, "Back 按钮可见"])
    steps_sheet.append(["DGX-002", 6, "assert_visible", "role=button[name='Next']", "", 3000, "Next 按钮可见"])
    steps_sheet.append(["DGX-002", 7, "screenshot", "", "header-controls", 3000, "保存导航区截图"])

    steps_sheet.append(["DGX-003", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"])
    steps_sheet.append(["DGX-003", 2, "wait_for_timeout", "", "2500", 2500, "等待地图渲染"])
    steps_sheet.append(["DGX-003", 3, "assert_visible", ".leaflet-container", "", 3000, "Leaflet 地图容器可见"])
    steps_sheet.append(["DGX-003", 4, "assert_visible", "text=Leaflet", "", 3000, "Leaflet 标识可见"])
    steps_sheet.append(["DGX-003", 5, "assert_visible", "text=OpenStreetMap contributors", "", 3000, "地图版权文字可见"])
    steps_sheet.append(["DGX-003", 6, "screenshot", "", "map-anchors", 3000, "保存地图锚点截图"])

    steps_sheet.append(["DGX-004", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"])
    steps_sheet.append(["DGX-004", 2, "wait_for_timeout", "", "2000", 2000, "等待前端挂载完成"])
    steps_sheet.append(["DGX-004", 3, "assert_count_gte", "body *", "5", 3000, "确认页面不是空白壳子"])
    steps_sheet.append(["DGX-004", 4, "screenshot", "", "dom-rendered", 3000, "保存渲染态截图"])

    steps_sheet.append(["DGX-005", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"])
    steps_sheet.append(["DGX-005", 2, "wait_for_timeout", "", "2500", 2500, "等待控制台稳定"])
    steps_sheet.append(["DGX-005", 3, "assert_page_errors_lte", "", "0", 3000, "页面脚本异常数量不超过阈值"])
    steps_sheet.append(["DGX-005", 4, "assert_console_errors_lte", "", "0", 3000, "控制台 error 数量不超过阈值"])

    steps_sheet.append(["DGX-006", 1, "goto", "", TARGET_URL, 15000, "打开目标页面"])
    steps_sheet.append(["DGX-006", 2, "wait_for_timeout", "", "2500", 2500, "等待地图文字渲染"])
    steps_sheet.append(["DGX-006", 3, "assert_visible", "text=Ferrari World Yas Island, Abu Dhabi", "", 5000, "目标地标文字可见"])

    guide_sheet = workbook.create_sheet("action_guide")
    guide_sheet.append(["action", "locator", "value", "说明"])
    guide_sheet.append(["goto", "", "URL", "打开页面"])
    guide_sheet.append(["wait_for_load_state", "", "load|domcontentloaded|networkidle", "等待页面状态"])
    guide_sheet.append(["wait_for_selector", "Playwright locator", "visible|hidden|attached|detached", "等待某个元素进入指定状态"])
    guide_sheet.append(["wait_for_timeout", "", "毫秒", "固定等待"])
    guide_sheet.append(["click", "Playwright locator", "", "点击元素"])
    guide_sheet.append(["fill", "Playwright locator", "文本", "输入文本"])
    guide_sheet.append(["press", "Playwright locator", "Enter", "键盘输入"])
    guide_sheet.append(["hover", "Playwright locator", "", "悬停元素"])
    guide_sheet.append(["select_option", "Playwright locator", "value", "下拉选择"])
    guide_sheet.append(["assert_title_contains", "", "文本", "断言标题包含文本"])
    guide_sheet.append(["assert_url_contains", "", "文本", "断言 URL 包含文本"])
    guide_sheet.append(["assert_visible", "Playwright locator", "", "断言元素可见"])
    guide_sheet.append(["assert_hidden", "Playwright locator", "", "断言元素隐藏"])
    guide_sheet.append(["assert_text_contains", "Playwright locator", "文本", "断言元素文本包含文本"])
    guide_sheet.append(["assert_count_gte", "Playwright locator", "整数", "断言匹配元素数量至少 N 个"])
    guide_sheet.append(["assert_page_errors_lte", "", "整数", "页面异常数量不超过阈值"])
    guide_sheet.append(["assert_console_errors_lte", "", "整数", "控制台 error 数量不超过阈值"])
    guide_sheet.append(["assert_request_failures_lte", "", "整数", "失败请求数量不超过阈值"])
    guide_sheet.append(["screenshot", "", "文件名", "截图保存到当前 case 目录"])
    guide_sheet.append(["", "", "", "locator 可以直接写 Playwright 选择器，例如 text=Leaflet、role=button[name='Next']、.leaflet-container"])

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Excel 模板已生成: {OUTPUT_PATH.resolve()}")


if __name__ == "__main__":
    main()
